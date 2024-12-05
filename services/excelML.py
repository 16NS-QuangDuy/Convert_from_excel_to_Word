# encoding: utf-8
import openpyxl
from openpyxl.utils import get_column_letter
import openpyxl.worksheet.merge
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, PatternFill, colors, Font
import re
from copy import copy


class ExcelML:
    """handler for doing excel operation"""

    def __init__(self, filename=None):
        self.__auto_worker_name = self.__class__.__name__
        self.filename = filename
        if filename:
            self.wb = openpyxl.load_workbook(self.filename)

    def get_ws(self, ws_name):
        ws = None
        try:
            ws = self.wb[ws_name]
        except:
            print ('- error: there is no worksheet name ', ws_name)
            pass
        return ws

    @staticmethod
    def copy_cell_range(sheet, sheet_dst, min_col, min_row, max_col, max_row, shift_col, shift_row):
        # Copy all cells
        for col in range( min_col, max_col+1):
            for row in range( min_row+1, max_row+1):
                fmt = "{min_col}{min_row}"
                # Create the copy source code.
                copySrcCoord = fmt.format(
                    min_col = get_column_letter(col),
                    min_row = row );
                # Create copy destination code.
                copyDstCoord = fmt.format(
                    min_col = get_column_letter(col + shift_col),
                    min_row = row + shift_row );
                # Copy the value to the destination.
                if type(sheet[copySrcCoord]) != MergedCell :
                    # If there is a format, copy the format.
                    if sheet[copySrcCoord].has_style :
                        # sheet[copyDstCoord]._style = sheet[copySrcCoord]._style;
                        sheet_dst[copyDstCoord]._style = sheet[copySrcCoord]._style;
        return 0

    @staticmethod
    def read_all_rows_from_excel_sheet(ws):
        start_row, from_col, width = ExcelML.get_true_start_row_from_column(ws)
        row_idx = 0
        records = []
        for row in ws.iter_rows():
            #rc = [cell.value for cell in row]
            rc = [ExcelFormatRule.check_format_and_get_value(cell).replace('_x000D_', '').strip() if isinstance(cell.value,
                                                                                                        str) else ExcelFormatRule.check_format_and_get_value(
                cell) for cell in row]
            if row_idx >= start_row:
                if rc[from_col] is None or rc[from_col] == "":
                    break
                else:
                    records.append(rc)
            row_idx += 1
        return records

    @staticmethod
    def read_all_rows_from_excel_sheet2(ws):
        start_row, from_col, width = ExcelML.get_true_start_row_from_column(ws)
        row_idx = 0
        records = []
        for row in ws.iter_rows():
            rc = [cell.value for cell in row]
            if row_idx >= start_row:
                records.append(rc)
            row_idx += 1
        return records

    @staticmethod
    def read_header_row_from_excel_sheet(ws, start_row=0):
        start_row, from_col, width = ExcelML.get_start_row_from_column(ws)
        row_idx = 0
        records = []
        for row in ws.iter_rows():
            if row_idx >= start_row:
                rc = [cell.value for cell in row]
                records.append(rc)
                break
            row_idx += 1
        return records

    @staticmethod
    def read_all_rows_with_gap_from_excel_sheet(ws, start, gap,min_column = None, max_column = None):
        records = []
        gap_count = 0
        default_range = 2
        mincol = min_column
        maxcol = max_column
        for row in ws.iter_rows(min_row=start, min_col = mincol, max_col = maxcol):
            blank_row = True
            if (min_column is None or max_column is None):
                range_check = default_range
            else:
                range_check = max_column - min_column
            for i in range(range_check): # check blank_row data
                b_value = row[i].value
                if b_value is not None:
                    blank_row = False
                    break
            if blank_row: # check blank gap
                gap_count = gap_count + 1
                if gap_count >= gap:
                    break
                continue
            else:
                gap_count = 0
            rc = []
            for cell in row:
                rc.append(cell.value)
            records.append(rc)
        return records

    @staticmethod
    def read_color_all_rows_from_excel_sheet(ws, start, key_col=1):
        records = []
        for row in ws.iter_rows(min_row=start):
            value = row[key_col].value
            if value is None or value == "":
                break
            rc = []
            for cell in row:
                rgb = None
                if cell.font.color:
                    rgb = cell.font.color.rgb
                rc.append(rgb)
            records.append(rc)
        return records

    @staticmethod
    def fill_yaml_text_to_cell(ws, row, column, input_yaml):
        if input_yaml is None:
            input_yaml = ""
        else:
            input_yaml = "%s" % input_yaml
        if input_yaml == "-" or input_yaml == "¾":
            input_yaml = "—"
        if input_yaml == "":
            return 0
        ws.cell(row, column).value = input_yaml
        ws.cell(row, column).alignment = Alignment(wrap_text=True)
        return 1

    @staticmethod
    def fill_text_to_cell(ws, row, column, input_text):
        if input_text is None:
            input_text = ""
        else:
            input_text = "%s" % input_text
        if input_text == "-" or input_text == "¾":
            input_text = "—"
        if input_text == "":
            return 0
        try:
            ws.cell(row, column).value = input_text
            ws.cell(row, column).alignment = Alignment(wrap_text=True)
        except:
            pass
        return 1

    def fill_text_to_cell2(ws, row, column, input_text):
        if input_text is None:
            input_text = ""
        else:
            input_text = "%s" % input_text
        if input_text == "-" or input_text == "¾":
            input_text = "—"
        if input_text == "":
            return 0
        try:
            input_text = int(input_text)
        except:
            pass
        try:
            ws.cell(row, column).value = input_text
            ws.cell(row, column).alignment = Alignment(wrap_text=True)
        except:
            pass
        return 1

    @staticmethod
    def read_record_field(rc_dict, rc_key):
        if rc_key in rc_dict:
            value = rc_dict[rc_key]
            if value is None:
                value = ""
        else:
            value = ""
        return value

    @staticmethod
    def read_excel_record_field(record, field_name, columns_idx_map):
        idx = columns_idx_map[field_name]
        if idx >= len(record):
            value = None
        else:
            value = record[columns_idx_map[field_name]]
        if value is None:
            value = ""
        elif isinstance(value, int):
            value = "%s" % value
        return value

    @staticmethod
    def get_start_row_from_column(ws, max=5):
        from_col = 0
        start_row = 0
        width = 0
        count = 0
        # find start row
        header = []
        for row in ws.iter_rows():
            rc = [cell.value for cell in row]
            rc = [str(rc_i).strip() for rc_i in rc if rc_i is not None]
            if rc.count("") >= len(rc) and count < max:
                start_row += 1
            else:
                header = [cell.value for cell in row]
                width = len(header)
                break
        # find from col
        for value in header:
            if value is None:
                from_col += 1
            else:
                break
        return start_row, from_col, width

    @staticmethod
    def get_true_start_row_from_column(ws, max=5):
        start_row, from_col, width = ExcelML.get_start_row_from_column(ws, max)
        # over the merge cell
        start_body = 0
        for row in ws.iter_rows():
            if start_body <= start_row:
                rc1 = [cell.value for cell in row]
            else:
                rc = [cell.value for cell in row]
                if rc[from_col] is not None or str(rc[from_col]).strip() == "":
                    break
            start_body += 1
        return start_body, from_col, width

    @staticmethod
    def get_width_header(row, from_col, max_width=30):
        width = -1
        for width in range(from_col+1, max_width):
            val = row[width].value
            if val is None:
                return width
        return width

    @staticmethod
    def delete_duplicate_rows_from_sheet(records, key_col_list):
        out_records = []
        for row in records:
            found = False
            for out_row in out_records:
                if ExcelML.is_duplicate_rows_no_test(row, out_row, key_col_list):
                    found = True
                    break
            if found:
                continue
            out_records.append(row)
        return out_records

    @staticmethod
    def is_duplicate_records_no_test(records_a, records_b, key_col_list=[]):
        if len(records_a) != len(records_b):
            return False
        else:
            match = 0
            checked_idx_list = []
            for row_a in records_a:
                found = False
                for idx, row_b in enumerate(records_b):
                    if idx in checked_idx_list:
                        continue
                    if ExcelML.is_duplicate_rows_no_test(row_a, row_b, key_col_list):
                        found = True
                        checked_idx_list.append(idx)
                        break
                if found:
                    match += 1
                    continue
            if match == len(records_a):
                return True
            else:
                return False

    @staticmethod
    def is_duplicate_rows_no_test(row_a, row_b, key_col_list=[]):
        if len(key_col_list) == 0:
            key_col_list = [i for i in range(len(row_a))]
        for idx in range(len(row_a)):
            if idx in key_col_list and row_a[idx] != row_b[idx]:
                return False
        return True

    @staticmethod
    def diff_excel_a_b(excel_a, excel_b, sheet_name, key_col_list):
        # 2 excel file have same template
        wk1 = openpyxl.load_workbook(excel_a)
        wk2 = openpyxl.load_workbook(excel_b)
        sht1 = ExcelML.get_sht_name(wk1, sheet_name)
        sht2 = ExcelML.get_sht_name(wk2, sheet_name)
        records1 = ExcelML.read_all_rows_from_excel_sheet(wk1[sht1])
        records2 = ExcelML.read_all_rows_from_excel_sheet(wk2[sht2])
        if len(records1) == len(records2) == 0:
            return {'result': True, 'diff_list_a': [], "diff_list_b": []}
        if len(records1) == 0 and len(records2) != 0:
            return {'result': False, 'diff_list_a': [sheet_name + "Empty"], "diff_list_b": [sheet_name + "Not Empty"]}
        if len(records1) != 0 and len(records2) == 0:
            return {'result': False, 'diff_list_a': [sheet_name + "Not Empty"], "diff_list_b": [sheet_name + "Empty"]}
        diff_list_a = []
        diff_list_b = []
        if len(records1) != len(records2):
            diff_list_a.append("NoOfRow A: %s" % len(records1))
            diff_list_b.append("NoOfRow B: %s" % len(records2))
        if len(records1[0]) != len(records2[0]):
            diff_list_a.append("NoOfColumn A: %s" % len(records1[0]))
            diff_list_b.append("NoOfColumn B: %s" % len(records2[0]))
        match = 0
        min_row = min(len(records1), len(records2))
        min_column = min(len(records1[0]), len(records2[0]))
        for i in range(min_row):
            for j in range(min_column):
                if key_col_list != [] and j not in key_col_list:
                    continue
                if records1[i][j] == records2[i][j]:
                    match += 1
                else:
                    cell_name = ExcelML.convert_cell_name(i, j)
                    diff_list_a.append(cell_name + "A: %s" % records1[i][j])
                    diff_list_b.append(cell_name + "B: %s" % records2[i][j])
        if len(diff_list_a) > 0:
            return {'result': False, 'diff_list_a': diff_list_a, "diff_list_b": diff_list_b}
        else:
            return {'result': True, 'diff_list_a': diff_list_a, "diff_list_b": diff_list_b}

    @staticmethod
    def diff_excel_a_b_all(excel_a, excel_b):
        # 2 excel file have same template
        wk1 = openpyxl.load_workbook(excel_a)
        sheet_list = wk1.sheetnames
        match = 0
        diff_list_a = []
        diff_list_b = []
        for sheet_name in sheet_list:
            result = ExcelML.diff_excel_a_b(excel_a, excel_b, sheet_name, [])
            if result["result"]:
                match += 1
            else:
                diff_list_a.extend(["%s %s" % (sheet_name, diff) for diff in result["diff_list_a"]])
                diff_list_b.extend(["%s %s" % (sheet_name, diff) for diff in result["diff_list_b"]])
        if match > 0 and len(diff_list_a) == len(diff_list_b) == 0:
            return {'result': True, 'diff_list_a': diff_list_a, "diff_list_b": diff_list_b}
        else:
            return {'result': False, 'diff_list_a': diff_list_a, "diff_list_b": diff_list_b}

    @staticmethod
    def convert_cell_name(row, column_idx):
        column_idx = chr(column_idx + ord('A'))
        return "CELL(%s%s)" % (column_idx, row, )

    @staticmethod
    def convert_cell_name2(row, column_idx):
        column_idx = chr(column_idx + ord('A'))
        return "%s%s" % (column_idx, row, )

    @staticmethod
    def copy_cells(source_sheet, target_sheet):
        for (row, col), source_cell in source_sheet._cells.items():
            target_cell = target_sheet.cell(column=col, row=row)

            target_cell._value = source_cell._value
            target_cell.data_type = source_cell.data_type

            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)

            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)

    @staticmethod
    def copy_sheet(source_sheet, target_sheet):
        ExcelML.copy_cells(source_sheet, target_sheet)  # copy all the cel values and styles
        ExcelML.copy_sheet_attributes(source_sheet, target_sheet)

    @staticmethod
    def copy_sheet_attributes(source_sheet, target_sheet):
        target_sheet.sheet_format = copy(source_sheet.sheet_format)
        target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
        target_sheet.merged_cells = copy(source_sheet.merged_cells)
        target_sheet.page_margins = copy(source_sheet.page_margins)
        target_sheet.freeze_panes = copy(source_sheet.freeze_panes)

        # set row dimensions
        # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
        for rn in range(len(source_sheet.row_dimensions)):
            target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])

        if source_sheet.sheet_format.defaultColWidth is not None:
            target_sheet.sheet_format.defaultColWidth = copy(source_sheet.sheet_format.defaultColWidth)

        # set specific column width and hidden property
        # we cannot copy the entire column_dimensions attribute so we copy selected attributes
        for key, value in source_sheet.column_dimensions.items():
            target_sheet.column_dimensions[key].min = copy(source_sheet.column_dimensions[
                                                               key].min)  # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
            target_sheet.column_dimensions[key].max = copy(source_sheet.column_dimensions[
                                                               key].max)  # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not onl;y the hidden property
            target_sheet.column_dimensions[key].width = copy(
                source_sheet.column_dimensions[key].width)  # set width for every column
            target_sheet.column_dimensions[key].hidden = copy(source_sheet.column_dimensions[key].hidden)

    @staticmethod
    def get_column_witdh(ws):
        """
        get data of column width of ws
        :param ws:
        :return: list of column width from 0
        """
        dims = {}
        for row in ws.rows:
            for cell in row:
                dims[cell.column] = ws.column_dimensions[openpyxl.utils.get_column_letter(cell.column)].width
            break
        data = [value for key, value in dims.items()]
        return data

    @staticmethod
    def set_columnn_width(ws, Width, data):
        """
        set column width of sheet ws with width and data dictionary
        :param ws:
        :param Width:
        :param data:
        :return:
        """
        for col in range(min(Width, len(data)-1)):
            width = data[col]
            ws.column_dimensions[openpyxl.utils.get_column_letter(col+1)].width = width

    @staticmethod
    def get_sht_name(wb, SheetName):
        """
        get sheet name from input name in case has space or no space name are both acceptable names
        :param wb:
        :param SheetName:
        :return:
        """
        found = False
        if isinstance(SheetName, str) and SheetName.count(" ") < 1:
            word_list = [match.group(0) for match in re.finditer(r"[A-Z][a-z]+", SheetName)]
            new_name = " ".join(word_list)
            old_name = SheetName
            SheetName = [old_name, new_name]
        if isinstance(SheetName, list):
            for sht in SheetName:
                if sht in wb.sheetnames:
                    actual_name = sht
                    found = True
                    break
            if not found:
                actual_name = SheetName
        else:
            actual_name = SheetName
        return actual_name

    @staticmethod
    def create_column_dict_from_headers(headers, is_lower=True):
        """
        create dictionary from column field and index
        :param headers: header records of fields
        :param is_lower: only lower key
        :return: dictionary
        """
        column_dict = dict()
        for header in headers:
            for idx, field in enumerate(header):
                if field is not None and field != "":
                    field = str(field)
                    if "[internal only]" in field.lower() and "\n" in field:
                        field = " ".join(field.split("\n"))
                        field = field.replace("nly] ", "nly]")
                    true_field = field if "\n" not in field else field.split("\n")[0]
                    true_field = true_field.strip()
                    if is_lower:
                        true_field = true_field.lower()
                        column_dict[true_field] = idx
                    else:
                        column_dict[true_field] = idx
                        true_field2 = true_field.lower()
                        column_dict[true_field2] = idx
                        word_list = true_field.split(" ")
                        for j, word in enumerate(word_list):
                            word_list[j] = word[:1].upper() + word[1:]
                        true_field3 = " ".join(word_list)
                        column_dict[true_field3] = idx
                        if true_field in ["Fusa", "FuSa", "fusa"]:
                            column_dict["Fusa"] = idx
                        true_field4 = true_field.lower()
                        true_field4 = true_field4[:1].upper() + true_field4[1:]
                        column_dict[true_field4] = idx
                        if true_field2 in ["Bit start"]:
                            pass

        return column_dict


class ExcelFormatRule:

    @staticmethod
    def organize_format_base_address_name(base_address_list):
        if len(base_address_list) > 1:
            out_list = []
            for idx, (base, address) in enumerate(base_address_list):
                base = "%s[%s]" % (base, idx)
                address = address.replace("0x", "H'")
                out_list.append((base, address))
            return out_list
        else:
            return [(base, address.replace("0x", "H'")) for (base, address) in base_address_list]

    @staticmethod
    def organize_format_address_offset(address_offset):
        return address_offset.replace("0x", "H'")

    @staticmethod
    def organize_configuration_address(base_address, address_offset, isArray=False):
        if isArray:
            return "<%s[n]> + %s" % (base_address, address_offset.replace("0x", "H'"))
        else:
            return "<%s> + %s" % (base_address, address_offset.replace("0x", "H'"))

    @staticmethod
    def organize_bit_name(bit_name):
        """
        this function is to do format the bit name in TS
        :param bit_name: bit name
        :return: well-form bit name
        """
        if bit_name in ["Reserved", "—", ""]:
            bit_name = "—"
        return bit_name

    @staticmethod
    def organize_bit_field(bit_list, access_size=32):
        """
        :param bit_list: list of bitname, width
        :param access_size: 32 as default
        :return: None
        """
        out_list = []
        start_bit, end_bit = access_size-1, 0
        for idx, (bitname, width) in enumerate(bit_list):
            end_bit = start_bit + 1 - int(width)
            out_list.append([bitname, "%s" % start_bit, "%s" % end_bit])
            start_bit = end_bit-1
        return out_list

    @staticmethod
    def organize_description_base_address(base_address, isArray=False):
        if isArray:
            return "<%s[n]>" % base_address
        else:
            return "<%s>" % base_address

    @staticmethod
    def check_format_and_get_value(cell):
        prefix_list = ['B', 'H']
        prefix_string = '|'.join(prefix_list)
        regex = re.compile(r'(\"({})\'\")(.)'.format(prefix_string))
        format = cell.number_format
        match = regex.search(format)
        if match and str(cell.value).isdigit():
            return format[format.find('"') + 1:format.rfind('"')] + str(cell.value)
        return cell.value

    @staticmethod
    def fill_cell_color_error(cell):
        cell.fill = PatternFill(start_color=colors.COLOR_INDEX[5], end_color=colors.COLOR_INDEX[5], fill_type='solid')
        # cell.alignment = Alignment(wrap_text)
        # cell.font = Font(size=11, color=colors.WHITE)

    @staticmethod
    def fill_cell_color_ok(cell):
        cell.fill = PatternFill(start_color=colors.COLOR_INDEX[3], end_color=colors.COLOR_INDEX[3], fill_type='solid')
        # cell.alignment = Alignment(wrap_text=True)

    @staticmethod
    def fill_cell_color_header(cell):
        # cell.fill = PatternFill(start_color=colors.DARKYELLOW, end_color=colors.DARKYELLOW, fill_type='solid')
        cell.alignment = Alignment(wrap_text=True)
        # cell.font = Font(size=11, color=colors.WHITE)

    @staticmethod
    def wrap_text_cell(cell):
        cell.alignment = Alignment(wrap_text=True)

