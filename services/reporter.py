# encoding: utf-8
import yaml
import openpyxl
import openpyxl.worksheet.merge
from services.excelML import ExcelML


class Reporter:
    """handler for doing report in excel template creation"""

    def __init__(self, input_template, input_config, auto_task_name="JapaneseWorker"):
        self.auto_task_name = auto_task_name
        self.__target_spec_config_dict = {}
        with open(input_config, 'r', encoding="utf-8") as stream:
            self.__target_spec_config_dict = yaml.safe_load(stream)
        self.detail_sheet = self.__target_spec_config_dict[auto_task_name]["SheetName"]
        self.column_list = self.__target_spec_config_dict[auto_task_name]["ColumnList"]
        self.column_dict = self.__target_spec_config_dict[auto_task_name]["ColumnDict"]
        self.start = self.__target_spec_config_dict[auto_task_name]["Start"]
        self.width = self.__target_spec_config_dict[auto_task_name]["Width"]
        self.header = self.start - 1
        self.__wb = openpyxl.load_workbook(input_template)
        self.__ws = self.__wb[self.detail_sheet]
        self.__records = []
        self.__row_offset = self.start
        self.__count = 0
        self.__sheet_list = []
        if "Header" in self.__target_spec_config_dict[auto_task_name]:
            self.header = self.__target_spec_config_dict[auto_task_name]["Header"]

    def create_detail_report(self, records):
        ws = self.__ws
        self.__row_offset = self.start
        row_offset = self.__row_offset
        for rc_id, rc in enumerate(records):
            row_idx = row_offset + rc_id
            # [rc_id, text, language, is_hidden, BlockID]
            for fld_idx, value in enumerate(rc):
                col_idx = fld_idx + 1
                ExcelML.fill_text_to_cell(ws, row_idx, col_idx, value)
            self.copy_or_append_one_template()
            self.__row_offset = row_idx

    def add_another_detail_report(self, input_config, records, auto_task_name):
        self.__target_spec_config_dict = {}
        with open(input_config, 'r', encoding="utf-8") as stream:
            self.__target_spec_config_dict = yaml.safe_load(stream)
        self.detail_sheet = self.__target_spec_config_dict[auto_task_name]["SheetName"]
        self.column_list = self.__target_spec_config_dict[auto_task_name]["ColumnList"]
        self.column_dict = self.__target_spec_config_dict[auto_task_name]["ColumnDict"]
        self.start = self.__target_spec_config_dict[auto_task_name]["Start"]
        self.width = self.__target_spec_config_dict[auto_task_name]["Width"]
        self.__ws = self.__wb[self.detail_sheet]
        self.create_detail_report(records)

    def add_another_detail_report2(self, sheet_name, records):
        self.detail_sheet = sheet_name
        self.__ws = self.__wb[self.detail_sheet]
        self.create_detail_report(records)

    def copy_or_append_one_template(self, size=1):
        source = self.__wb[self.detail_sheet]
        if self.__row_offset > self.start:
            ExcelML.copy_cell_range(
                sheet=source,
                sheet_dst=source,
                min_col=1,
                min_row=self.start,
                max_col=self.width,
                max_row=self.start + size,
                shift_col=0,
                shift_row=self.__row_offset
            );

    def read_titles_template(self, condition="", row_id=1):
        ws = self.__ws
        max_col = ws.max_column
        tickets = []
        for col in range(3, max_col):
            c = ws.cell(row_id, col)
            if c.value is None:
                break
            if c.value.startswith(condition):
                tickets.append(c.value)
        return set(tickets)

    def write_value_follow_title(self, post: dict, row_id=1):
        ws = self.__ws
        max_col = ws.max_column
        for col in range(3, max_col):
            c = ws.cell(row_id, col)
            content_key = post.get(c.value)
            if content_key is not None:
                w = ws.cell(row_id + 1, col)
                t = []
                for i, link in enumerate(content_key):
                    t.append(link)
                t = "\n ".join(t)
                w.value = t
        self.__ws = ws
        return self.__ws

    def save(self, target_xlsx):
        save_file_name = target_xlsx
        self.__wb.save(save_file_name)
        self.__wb.close()

    def create_detail_table_data(self, table_list, sht_name_lst=[]):
        self.__temp_sheet = self.__target_spec_config_dict[self.auto_task_name]["TemplateSheetName"]
        bk = self.detail_sheet
        for i, table in enumerate(table_list):
            lcount = i + 1
            record = table["record"]
            sheet_name = 'T' + "0" * (3 - len("%s" % lcount)) + '%d' % lcount if sht_name_lst == [] else sht_name_lst[i]
            source = self.__wb[self.__temp_sheet]
            ws = self.__wb.copy_worksheet(source)
            ws.title = sheet_name
            self.__row_offset = self.start
            row_offset = self.__row_offset
            self.detail_sheet = sheet_name
            for rc_id, rc in enumerate(record):
                row_idx = row_offset + rc_id
                # [rc_id, text, language, is_hidden, BlockID]
                for fld_idx, value in enumerate(rc):
                    col_idx = fld_idx + 1
                    ExcelML.fill_text_to_cell(ws, row_idx, col_idx,
                                              value)
                self.copy_or_append_one_template()
                self.__row_offset = row_idx
        self.detail_sheet = bk

    def create_link_summary_table_data(self, table_list, sht_name_lst=[]):
        self.__summary_sheet = self.detail_sheet
        col_link = self.column_dict["Link"]
        col_link = ord(col_link) - ord('A') + 1
        self.row_offset = self.start
        ws_summary = self.__wb[self.__summary_sheet]
        ExcelML.fill_yaml_text_to_cell(ws_summary, 1, 1, "HOME")
        for i, table in enumerate(table_list):
            table_title = table["table_title"]
            lcount = i + 1
            sheet_name = 'T' + "0" * (3 - len("%s" % lcount)) + '%d' % lcount if sht_name_lst == [] else sht_name_lst[i]
            ws = self.__wb[sheet_name]
            link = '=HYPERLINK(' + '"#%s!A1"' % self.__summary_sheet + ',"[Goto Summary]")'
            ExcelML.fill_yaml_text_to_cell(ws, 1, 1, link)
            ExcelML.fill_yaml_text_to_cell(ws, 1, 2, table_title)
            link_detail = '=HYPERLINK(' + '"#%s!A%s"' % (sheet_name, 2) + ',"For details")'
            row_idx = self.row_offset + i
            ExcelML.fill_yaml_text_to_cell(ws_summary, row_idx, col_link, link_detail)

    @staticmethod
    def create_command_window_report(auto_worker_name, data_list, data_label, data_type="ITEMS"):
        print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        print("      %s SUMMARY REPORT    " % auto_worker_name)
        print("      - NUMBER OF %s:" % data_type, len(data_list))
        print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        for text in data_list:
            try:
                print("      - %s             :" % data_label, text)
            except:
                pass
        print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")

    def update_headers(self, headers, header_row=None):
        ws = self.__ws
        row_offset = self.header if not header_row else header_row
        for rc_id, rc in enumerate(headers):
            row_idx = row_offset + rc_id
            for fld_idx, value in enumerate(rc):
                col_idx = fld_idx + 1
                ExcelML.fill_text_to_cell(ws, row_idx, col_idx, value)
            row_offset = row_idx

    def get_ws(self):
        return self.__wb

    def remove_sheet(self, shtname):
        self.__wb.remove_sheet(self.__wb[shtname])
