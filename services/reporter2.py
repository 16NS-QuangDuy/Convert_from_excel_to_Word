# encoding: utf-8
import yaml
import openpyxl
import openpyxl.worksheet.merge
from services.excelML import ExcelML


class Reporter2:
    """handler for doing report in excel template creation"""

    def __init__(self, input_template, config_dict=None):
        self.wb = openpyxl.load_workbook(input_template)
        if config_dict is None:
            return
        self.config_dict = config_dict
        self.detail_sheet = self.config_dict["SheetName"]
        self.start = self.config_dict["Start"]
        self.width = self.config_dict["Width"]
        self.header = self.start - 1
        self.ws = self.wb[self.detail_sheet]
        self.__records = []
        self.row_offset = self.start
        self.__count = 0
        self.__sheet_list = []
        if "Header" in self.config_dict:
            self.header = self.config_dict["Header"]

    def create_detail_report(self, records):
        ws = self.ws
        self.row_offset = self.start
        row_offset = self.row_offset
        for rc_id, rc in enumerate(records):
            row_idx = row_offset + rc_id
            # [rc_id, text, language, is_hidden, BlockID]
            for fld_idx, value in enumerate(rc):
                col_idx = fld_idx + 1
                ExcelML.fill_text_to_cell2(ws, row_idx, col_idx, value)
            self.copy_or_append_one_template()
            self.row_offset = row_idx

    def add_another_detail_report(self, config_dict, records):
        self.config_dict = config_dict
        self.detail_sheet = self.config_dict["SheetName"]
        self.start = self.config_dict["Start"]
        self.width = self.config_dict["Width"]
        self.detail_sheet = ExcelML.get_sht_name(self.wb, self.detail_sheet)
        self.ws = self.wb[self.detail_sheet]
        self.create_detail_report(records)

    def add_another_detail_report2(self, sheet_name, records):
        self.detail_sheet = sheet_name
        self.ws = self.wb[self.detail_sheet]
        self.create_detail_report(records)

    def copy_or_append_one_template(self, size=1):
        source = self.wb[self.detail_sheet]
        if self.row_offset > self.start:
            ExcelML.copy_cell_range(
                sheet=source,
                sheet_dst=source,
                min_col=1,
                min_row=self.start,
                max_col=self.width,
                max_row=self.start + size,
                shift_col=0,
                shift_row=self.row_offset
            );

    def save(self, target_xlsx):
        save_file_name = target_xlsx
        self.wb.save(save_file_name)
        self.wb.close()

    def create_detail_table_data(self, table_list):
        self.temp_sheet = self.config_dict["TemplateSheetName"]
        for i, table in enumerate(table_list):
            lcount = i + 1
            record = table["record"]
            sheet_name = 'T' + "0" * (3 - len("%s" % lcount)) + '%d' % lcount
            source = self.wb[self.temp_sheet]
            ws = self.wb.copy_worksheet(source)
            ws.title = sheet_name
            self.row_offset = self.start
            row_offset = self.row_offset
            for rc_id, rc in enumerate(record):
                row_idx = row_offset + rc_id
                # [rc_id, text, language, is_hidden, BlockID]
                for fld_idx, value in enumerate(rc):
                    col_idx = fld_idx + 1
                    ExcelML.fill_text_to_cell(ws, row_idx, col_idx,
                                                          value)
                self.copy_or_append_one_template()
                self.row_offset = row_idx

    def create_link_summary_table_data(self, table_list):
        self.summary_sheet = self.detail_sheet
        col_link = self.column_dict["Link"]
        col_link = ord(col_link) - ord('A') + 1
        self.row_offset = self.start
        ws_summary = self.wb[self.summary_sheet]
        ExcelML.fill_yaml_text_to_cell(ws_summary, 1, 1, "HOME")
        for i, table in enumerate(table_list):
            table_title = table["table_title"]
            lcount = i + 1
            sheet_name = 'T' + "0" * (3 - len("%s" % lcount)) + '%d' % lcount
            ws = self.wb[sheet_name]
            link = '=HYPERLINK(' + '"#%s!A1"' % self.summary_sheet + ',"[Goto Summary]")'
            ExcelML.fill_yaml_text_to_cell(ws, 1, 1, link)
            ExcelML.fill_yaml_text_to_cell(ws, 1, 2, table_title)
            link_detail = '=HYPERLINK(' + '"#%s!A%s"' % (sheet_name, 2) + ',"For details")'
            row_idx = self.row_offset + i
            ExcelML.fill_yaml_text_to_cell(ws_summary, row_idx, col_link, link_detail)

    @staticmethod
    def create_command_window_report(auto_worker_name, data_list, data_label, data_type="ITEMS"):
        print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        print("      %s SUMMARY REPORT    " % auto_worker_name)
        print("      - NUMBER OF %s:" % data_type, len(data_list))
        print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        for text in data_list:
            try:
                print("      - %s             :" % data_label, text)
            except:
                pass
        print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")

    def update_headers(self, headers, header_row=None):
        ws = self.ws
        self.header = 1
        row_offset = self.header if not header_row else header_row
        for rc_id, rc in enumerate(headers):
            row_idx = row_offset + rc_id
            for fld_idx, value in enumerate(rc):
                col_idx = fld_idx + 1
                current_value = ws.cell(row_idx, col_idx).value
                if current_value is None or current_value != value:
                    ExcelML.fill_text_to_cell(ws, row_idx, col_idx, value)
            row_offset = row_idx

    def set_columnn_width(self, data):
        ws = self.ws
        Width = self.width
        for col in range(min(Width, len(data)-1)):
            width = data[col]
            ws.column_dimensions[openpyxl.utils.get_column_letter(col+1)].width = width

    def get_wb(self):
        return self.wb

    def get_ws(self, sht):
        return self.wb[sht]

    def add_image(self, sht, cell, image_file):
        """
        add image file into cell
        JPG

        """
        img = openpyxl.drawing.image.Image(image_file)
        img.anchor = cell
        ws = self.wb[sht]
        ws.add_image(img)

